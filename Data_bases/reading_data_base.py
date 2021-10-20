# File to contain a function, that will read a data from exel sheets.

# ============ Dictionaries ============#
from openpyxl import load_workbook


# Reading data about player character
def load_data_base(file_path):

    # creating list to contain data
    data_base = []

    #iterator
    i = 1

    # open file
    wb = load_workbook(file_path)  # function to open exel file
    sheet = wb.active

    # assigned data to list data_base [] by using for loop.
    # while sheet.cell() return True value if it is not empty
    while sheet.cell(row=i, column=1).value:
        data_base.append(sheet.cell(row=i, column=1).value)
        i += 1

    # return data base
    return data_base

